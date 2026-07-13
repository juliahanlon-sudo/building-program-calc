#!/usr/bin/env python3
"""Build an Excel template: every space type as a row, one column per tier for
the manager to fill in the default. Data mirrors SPACE_GROUPS / DEFAULT_TIERS in App.jsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TIERS = [
    ("T1", "Tier 1 — Towers (Top Customer Markets)"),
    ("T2", "Tier 2 — Towers (Customer Hub)"),
    ("T3", "Tier 3 — Hubs (Employee Hub)"),
    ("T4", "Tier 4 — Leased Office"),
    ("T5", "Tier 5 — Serviced Offices"),
]

# (super group, group, [(space label, default SF, capacity type), ...])
GROUPS = [
    ("Workspace", "Individual Work", [
        ("Desks", 50, "Capacity"),
        ("Private Office", 150, "Capacity"),
        ("Touchdown Seat", 36, "Capacity"),
        ("Library", 60, "Capacity"),
        ("Work Room", 100, "Capacity"),
    ]),
    ("Workspace", "Enclosed Collaboration", [
        ("Phone Room (Micro)", 25, "Non-capacity"),
        ("Focus Pod", 35, "Non-capacity"),
        ("Meeting Pod", 50, "Non-capacity"),
        ("Phone Room", 50, "Non-capacity"),
        ("Phone Room (AV)", 60, "Non-capacity"),
        ("Huddle Room", 120, "Non-capacity"),
        ("Conference Room (M)", 250, "Non-capacity"),
        ("Conference Room (L)", 400, "Non-capacity"),
        ("Conference Room (XL)", 600, "Non-capacity"),
        ("Conference Room (Aloha)", 400, "Non-capacity"),
    ]),
    ("Workspace", "Workplace Specialty", [
        ("Social Lounge", 400, "Non-capacity"),
        ("Water Point", 200, "None"),
        ("Mindfulness", 250, "Non-capacity"),
        ("Reflection Room", 500, "Non-capacity"),
        ("Catering Pantry", 150, "Non-capacity"),
        ("Flex Room", 200, "Non-capacity"),
        ("Multifaith Room", 150, "Non-capacity"),
        ("Parent's Room", 100, "Non-capacity"),
        ("Reception", 200, "Non-capacity"),
        ("Reception Lounge", 300, "Non-capacity"),
        ("Treadmill Desk", 40, "Non-capacity"),
        ("Webinar Room", 200, "Non-capacity"),
        ("Wellness Room", 350, "Non-capacity"),
    ]),
    ("Workspace", "Open Collaboration", [
        ("Booth", 80, "Capacity"),
        ("Cafe Collaboration Table", 60, "Capacity"),
        ("Collaboration Space", 120, "Capacity"),
        ("Community Table", 50, "Capacity"),
        ("Project Bay", 150, "Capacity"),
        ("Soft Seating", 40, "Non-capacity"),
    ]),
    ("Amenity", "M&E", [
        ("Auditorium", 2000, "Capacity"),
        ("Pre-Function Space", 1500, "Non-capacity"),
        ("Training Room (L)", 1000, "Capacity"),
        ("Training Room (M)", 700, "Capacity"),
        ("Training Room (S)", 500, "Capacity"),
        ("Project Room", 300, "Capacity"),
    ]),
    ("Amenity", "Building Specialty", [
        ("AV Control Room", 200, "Non-capacity"),
        ("Badge Room", 200, "Non-capacity"),
        ("Catering", 400, "None"),
        ("Childcare", 600, "None"),
        ("Command Center", 400, "Capacity"),
        ("Critical Incident Center", 300, "Capacity"),
        ("CSIRT", 300, "Capacity"),
        ("Fitness Center", 800, "None"),
        ("Game Room", 300, "Non-capacity"),
        ("Go Center", 200, "Capacity"),
        ("IT Provisioning", 150, "Non-capacity"),
        ("Lab", 400, "Non-capacity"),
        ("Media Room", 200, "Non-capacity"),
        ("Medical Room", 150, "Non-capacity"),
        ("Mobility Lab", 300, "Non-capacity"),
        ("Outdoor / Terrace", 500, "Non-capacity"),
        ("Pantry", 200, "None"),
        ("Site Reliability Eng.", 300, "Capacity"),
        ("Staging / Green Room", 100, "None"),
        ("Techforce", 200, "Non-capacity"),
        ("Techforce Lab", 300, "Non-capacity"),
        ("UX Lab", 100, "Non-capacity"),
    ]),
    ("Amenity", "Hospitality", [
        ("AI - Learning", 200, "Non-capacity"),
        ("Barista Bar", 150, "Non-capacity"),
        ("Cafeteria", 800, "Non-capacity"),
        ("Community Trailblazer Hub", 400, "Non-capacity"),
        ("Customer - Conference Room", 250, "Capacity"),
        ("Customer - Huddle Room", 120, "Non-capacity"),
        ("Customer - Phone Room", 50, "Non-capacity"),
        ("Customer Work Room", 150, "Capacity"),
        ("Demo Area", 300, "Non-capacity"),
        ("Ohana - Conference Room", 250, "Capacity"),
        ("Ohana - Exhibition Dining", 600, "Non-capacity"),
        ("Ohana - Exhibition Kitchen (FOH)", 400, "Non-capacity"),
        ("Ohana - Huddle Room", 500, "Non-capacity"),
        ("Lounge", 400, "Non-capacity"),
        ("Ohana - Piano", 100, "Non-capacity"),
        ("Ohana - Production Room", 200, "Non-capacity"),
        ("Salon", 200, "Non-capacity"),
        ("SIC - Conference Room", 250, "Non-capacity"),
        ("SIC - Phone Room", 50, "Non-capacity"),
        ("SIC Private Dining", 300, "Non-capacity"),
    ]),
    ("Support", "Support", [
        ("AV Rack Room", 200, "Non-capacity"),
        ("AV Storage", 200, "None"),
        ("Bike Room", 80, "None"),
        ("Bomb Shelter", 200, "None"),
        ("Built Out Zone", 200, "None"),
        ("Coat Closet", 50, "None"),
        ("Communicating Stair", 300, "None"),
        ("Copy Print Center", 80, "None"),
        ("Cubbies", 20, "None"),
        ("IDF", 100, "None"),
        ("Irrigation Room", 80, "None"),
        ("Janitor Closet", 50, "None"),
        ("Lobby", 300, "None"),
        ("Locker Room", 150, "None"),
        ("Mail Center", 150, "Non-capacity"),
        ("MDF", 150, "None"),
        ("Millwork & Trash/Recycling", 7, "None"),
        ("Office Services Supply Rm", 100, "None"),
        ("Restroom", 200, "None"),
        ("Staff Room", 100, "Non-capacity"),
        ("Storage", 150, "None"),
        ("Team Storage", 7, "None"),
    ]),
]

# ── Styling ──────────────────────────────────────────────────────────────
NAVY = "001E5B"
BLUE = "066AFE"
LIGHT = "EEF4FF"
GREY = "F2F2F2"
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Space Defaults by Tier"

# Title + instructions
ws.merge_cells("A1:I1")
ws["A1"] = "Salesforce Space Planning — Default Space Types by Tier"
ws["A1"].font = Font(size=15, bold=True, color=NAVY)
ws.merge_cells("A2:I2")
ws["A2"] = ("Fill in the default quantity (or ratio) for each space type under each tier. "
            "Leave blank if the space type is not used in that tier.")
ws["A2"].font = Font(size=10, italic=True, color="514F4D")

header_row = 4
headers = ["Super Group", "Space Group", "Space Type", "Typical SF", "Capacity Type"] + [t[0] for t in TIERS]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Tier full-name note row under the tier letters
tier_note_row = header_row  # already have letters; add comments
for i, (tid, tlabel) in enumerate(TIERS):
    col = 6 + i
    ws.cell(row=header_row, column=col).comment = None

r = header_row + 1
first_tier_col = 6
for super_group, group, spaces in GROUPS:
    for label, sf, cap in spaces:
        ws.cell(row=r, column=1, value=super_group)
        ws.cell(row=r, column=2, value=group)
        ws.cell(row=r, column=3, value=label)
        ws.cell(row=r, column=4, value=sf)
        ws.cell(row=r, column=5, value=cap)
        for i in range(len(TIERS)):
            tc = ws.cell(row=r, column=first_tier_col + i)
            tc.fill = PatternFill("solid", fgColor=LIGHT)
        # borders + alignment for the row
        for c in range(1, 6 + len(TIERS)):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c in (1, 2):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        r += 1

# Zebra / group shading on the first two columns to make groups readable
for row in range(header_row + 1, r):
    sg = ws.cell(row=row, column=1).value
    grp = ws.cell(row=row, column=2).value

# Column widths
widths = {"A": 14, "B": 22, "C": 34, "D": 11, "E": 15}
for i in range(len(TIERS)):
    widths[get_column_letter(first_tier_col + i)] = 9
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze panes: keep header + first 3 descriptor columns visible
ws.freeze_panes = "F5"

# ── Reference sheet: tier definitions ───────────────────────────────────
ref = wb.create_sheet("Tier Reference")
ref["A1"] = "Tier Definitions"
ref["A1"].font = Font(size=14, bold=True, color=NAVY)
ref_headers = ["Tier", "Name / Examples", "Workspace %", "Amenity %", "Support %", "Density (SF/seat)"]
for c, h in enumerate(ref_headers, start=1):
    cell = ref.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

tier_ref = [
    ("T1", "Tier 1 — Towers (Top Customer Markets). e.g. New York, London, San Francisco, Tokyo", "65%", "25%", "10%", "85–110"),
    ("T2", "Tier 2 — Towers (Customer Hub). e.g. Chicago, Sydney", "70%", "20%", "10%", "85–105"),
    ("T3", "Tier 3 — Hubs (Employee Hub). e.g. Hyderabad", "78%", "14%", "8%", "75–95"),
    ("T4", "Tier 4 — Leased Office. All other leased offices", "82%", "8%", "10%", "70–90"),
    ("T5", "Tier 5 — Serviced Offices. All e-suites", "90%", "0%", "10%", "70–85"),
]
rr = 4
for row in tier_ref:
    for c, val in enumerate(row, start=1):
        cell = ref.cell(row=rr, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="left" if c == 2 else "center", vertical="center")
    rr += 1
ref.column_dimensions["A"].width = 8
ref.column_dimensions["B"].width = 60
for col in ("C", "D", "E", "F"):
    ref.column_dimensions[col].width = 15

out = "/Users/julia.hanlon/Desktop/Space-Types-by-Tier-Template.xlsx"
wb.save(out)
print("Saved:", out)
print("Space type rows:", sum(len(s) for _, _, s in GROUPS))
